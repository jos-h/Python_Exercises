from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import csv

def csv_Excel_conversion(csv_input_file):

	
	workbook = Workbook()

	# new workbook created with atleast one Worksheet
	worksheet = workbook.worksheets[0]
	worksheet.title = "a demo "
	
	# a variable to separate csv based on ','
	csv_separator = ','
	
	''' 
		register dialect as comma
		it can be anything
		ex:- :,|,# etc
	'''
	csv.register_dialect('comma', delimiter=',')
	
	'''
		itertating throughout the csv file
		and opening the file in read mode

	'''
	with open(csv_input_file,"r") as csv_file:
		
		'''
			will iterate through the csv file and 
			return a string it's next() method is executed
		'''
		reader = csv.reader(csv_file, dialect='comma')
		
		'''
			enumerate returns value in tuple form
			(a,b)
		'''
		for row_index, rows in enumerate(reader):

			for c, cols in enumerate(rows):
				
				for index, values in enumerate(cols.split(csv_separator)):

					'''
						ws.cell()
						gets us the content from the specified cells
					'''
					cell = worksheet.cell(row = row_index + 1, column = c + 1)
					cell.value = values
					index += 1
		'''
			OR

			#for column_index, cell in enumerate(rows):
			#column_letter = get_column_letter((column_index + 1 ))
			#worksheet.cell('%s%s'%(column_letter, (row_index + 1))).value = cell
			
		'''
	workbook.save("D://macrocodesrequired//sample.xlsx")


def main():
	
	
	csv_Excel_conversion("D://macrocodesrequired//SampleCSVFile_119kb.csv")


if __name__ == '__main__':
	main()